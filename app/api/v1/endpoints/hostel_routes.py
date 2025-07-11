# 🚀 FastAPI Router for Hostel Management

from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from datetime import datetime
from bson import ObjectId
from io import BytesIO
import openpyxl
from app.models.hostel import *  # Assuming models moved to app.models.hostel

hostelRouter = APIRouter()

# ---------- HOSTEL ----------
@hostelRouter.post("/create")
def create_hostel(data: dict):
    hostel = Hostel(**data)
    hostel.save()
    return str(hostel.id)

@hostelRouter.get("/list")
def list_hostels(school_id: str):
    return [h.to_mongo().to_dict() for h in Hostel.objects(school_id=school_id)]

@hostelRouter.put("/update/{hostel_id}")
def update_hostel(hostel_id: str, data: dict):
    h = Hostel.objects(id=hostel_id).first()
    if not h:
        raise HTTPException(404)
    h.update(**data)
    return "Updated"

@hostelRouter.delete("/delete/{hostel_id}")
def delete_hostel(hostel_id: str):
    h = Hostel.objects(id=hostel_id).first()
    if not h:
        raise HTTPException(404)
    h.delete()
    return "Deleted"

# ---------- BLOCK ----------
@hostelRouter.post("/block")
def create_block(data: dict):
    block = Block(**data)
    block.save()
    return str(block.id)

@hostelRouter.get("/blocks/{hostel_id}")
def get_blocks(hostel_id: str):
    return [b.to_mongo().to_dict() for b in Block.objects(hostel=hostel_id)]

# ---------- ROOM ----------
@hostelRouter.post("/room")
def create_room(data: dict):
    room = HostelRoom(**data)
    room.save()
    return str(room.id)

@hostelRouter.get("/rooms/{block_id}")
def list_rooms(block_id: str):
    return [r.to_mongo().to_dict() for r in HostelRoom.objects(block=block_id)]

# ---------- ALLOCATION ----------
@hostelRouter.post("/allocate")
def allocate_room(data: dict):
    student_id = data["student"]
    room = HostelRoom.objects(id=data["room"]).first()
    if not room:
        raise HTTPException(404, "Room not found")
    if room.occupied >= room.capacity:
        raise HTTPException(400, "Room Full")
    allocation = HostelAllocation(**data)
    allocation.save()
    room.occupied += 1
    room.save()
    return str(allocation.id)

@hostelRouter.get("/allocation/{student_id}")
def get_allocation(student_id: str):
    a = HostelAllocation.objects(student=student_id).first()
    if not a:
        raise HTTPException(404)
    return a.to_mongo().to_dict()

@hostelRouter.put("/deallocate/{student_id}")
def deallocate(student_id: str):
    a = HostelAllocation.objects(student=student_id).first()
    if not a:
        raise HTTPException(404)
    room = a.room
    if room: room.update(dec__occupied=1)
    a.left_at = datetime.utcnow()
    a.save()
    return "Deallocated"

# ---------- BULK UPLOAD ----------
@hostelRouter.post("/upload-allocations")
def bulk_allocate(school_id: str, file: UploadFile = File(...)):
    wb = openpyxl.load_workbook(file.file)
    sheet = wb.active
    added = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        student_id, hostel_id, block_id, room_id = row
        room = HostelRoom.objects(id=room_id).first()
        if room and room.occupied < room.capacity:
            alloc = HostelAllocation(
                student=student_id,
                hostel=hostel_id,
                block=block_id,
                room=room_id
            )
            alloc.save()
            room.occupied += 1
            room.save()
            added.append(student_id)
    return {"allocated": added}

@hostelRouter.get("/download/sample-upload")
def download_sample_excel():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["student_id", "hostel_id", "block_id", "room_id"])
    sheet.append(["student123", "hostel123", "block123", "room123"])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": "attachment; filename=sample_hostel_allocation.xlsx"})

# ---------- REPORTS ----------
@hostelRouter.get("/report/room-occupancy")
def room_occupancy_report():
    rooms = HostelRoom.objects()
    return [{"room": r.room_number, "capacity": r.capacity, "occupied": r.occupied} for r in rooms]

@hostelRouter.get("/report/hostel-capacity")
def hostel_capacity():
    hostels = Hostel.objects()
    result = []
    for h in hostels:
        rooms = HostelRoom.objects(block__hostel=h.id)
        total = sum(r.capacity for r in rooms)
        used = sum(r.occupied for r in rooms)
        result.append({"hostel": h.name, "total": total, "used": used})
    return result

@hostelRouter.get("/report/vacate-list")
def vacate_list():
    soon_to_vacate = HostelAllocation.objects(left_at__exists=True)
    return [{"student": str(a.student.id), "room": str(a.room.id), "left_at": a.left_at} for a in soon_to_vacate]

@hostelRouter.get("/export/room-occupancy")
def export_room_occupancy():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Room No", "Capacity", "Occupied"])
    for r in HostelRoom.objects():
        sheet.append([r.room_number, r.capacity, r.occupied])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": "attachment; filename=room_occupancy_report.xlsx"})
