
from datetime import datetime
from typing import Optional
from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
import openpyxl
import qrcode
from app.models.inventory import Asset, AssetMaintenance, AssetMovement, Category
from app.schema.inventroy_schema import AssetIn, CategoryIn, MaintenanceIn, MoveAssetIn
from fastapi.responses import StreamingResponse
from io import BytesIO
inventoryRouter = APIRouter()
@inventoryRouter.post("/category")
def add_category(data: CategoryIn):
    cat = Category(**data.dict())
    cat.save()
    return str(cat.id)

@inventoryRouter.get("/categories")
def get_categories(school_id: str):
    return [{"id": str(c.id), "name": c.name} for c in Category.objects(school_id=school_id)]

@inventoryRouter.delete("/category/{cat_id}")
def delete_category(cat_id: str):
    cat = Category.objects(id=cat_id).first()
    if not cat:
        raise HTTPException(404)
    cat.delete()
    return "deleted"

# ASSET ENDPOINTS
@inventoryRouter.post("/asset")
def add_asset(data: AssetIn):
    asset = Asset(**data.dict())
    asset.save()
    return str(asset.id)

@inventoryRouter.get("/assets")
def list_assets(school_id: str, category_id: Optional[str] = None):
    query = Asset.objects(school_id=school_id, is_deleted=False)
    if category_id:
        query = query.filter(category=category_id)
    return [{"id": str(a.id), "name": a.name, "location": a.location, "qty": a.quantity} for a in query]

@inventoryRouter.get("/asset/{asset_id}")
def get_asset(asset_id: str):
    asset = Asset.objects(id=asset_id).first()
    if not asset:
        raise HTTPException(404)
    return asset

@inventoryRouter.put("/asset/{asset_id}")
def update_asset(asset_id: str, data: AssetIn):
    asset = Asset.objects(id=asset_id).first()
    if not asset:
        raise HTTPException(404)
    asset.update(**data.dict())
    return "updated"

@inventoryRouter.delete("/asset/{asset_id}")
def delete_asset(asset_id: str):
    asset = Asset.objects(id=asset_id).first()
    if not asset:
        raise HTTPException(404)
    asset.is_deleted = True
    asset.save()
    return "soft-deleted"

# MOVEMENT
@inventoryRouter.post("/move")
def move_asset(data: MoveAssetIn):
    asset = Asset.objects(id=data.asset_id).first()
    if not asset:
        raise HTTPException(404)
    move = AssetMovement(
        asset=asset,
        from_location=data.from_location,
        to_location=data.to_location,
        moved_by=data.moved_by
    )
    asset.location = data.to_location
    asset.save()
    move.save()
    return "moved"

@inventoryRouter.get("/movements/{asset_id}")
def movement_history(asset_id: str):
    logs = AssetMovement.objects(asset=asset_id)
    return [{"from": l.from_location, "to": l.to_location, "at": l.moved_at.isoformat()} for l in logs]

# MAINTENANCE
@inventoryRouter.post("/maintenance")
def log_maintenance(data: MaintenanceIn):
    asset = Asset.objects(id=data.asset_id).first()
    if not asset:
        raise HTTPException(404)
    m = AssetMaintenance(asset=asset, cost=data.cost, notes=data.notes)
    m.save()
    return str(m.id)

@inventoryRouter.get("/maintenance/{asset_id}")
def list_maintenance(asset_id: str):
    logs = AssetMaintenance.objects(asset=asset_id)
    return [{"status": l.status, "cost": l.cost, "notes": l.notes, "reported_at": l.reported_at.isoformat()} for l in logs]

@inventoryRouter.put("/maintenance/{maint_id}")
def complete_maintenance(maint_id: str):
    m = AssetMaintenance.objects(id=maint_id).first()
    if not m:
        raise HTTPException(404)
    m.status = "Completed"
    m.completed_at = datetime.utcnow()
    m.save()
    return "marked completed"

# EXCEL UPLOAD
@inventoryRouter.post("/assets/upload-excel")
def upload_assets_excel(school_id: str, file: UploadFile = File(...)):
    wb = openpyxl.load_workbook(file.file)
    sheet = wb.active
    added = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, category_name, qty, location, condition = row
        category = Category.objects(school_id=school_id, name=category_name).first()
        if not category:
            category = Category(school_id=school_id, name=category_name)
            category.save()
        asset = Asset(
            school_id=school_id,
            category=category,
            name=name,
            quantity=qty,
            location=location,
            condition=condition
        )
        asset.save()
        added.append(str(asset.id))
    return {"added": added}

@inventoryRouter.get("/assets/sample-excel")
def download_sample_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asset Template"
    ws.append(["Name", "Category", "Quantity", "Location", "Condition"])
    ws.append(["Whiteboard", "Furniture", 5, "Class 9-A", "Working"])
    ws.append(["Projector", "Electronics", 2, "AV Room", "Working"])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=asset_sample.xlsx"})

# QR CODE
@inventoryRouter.get("/asset/{asset_id}/qr")
def get_qr(asset_id: str):
    url = f"https://school-system.com/assets/{asset_id}"
    img = qrcode.make(url)
    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return StreamingResponse(buf, media_type="image/png")

# REPORTS
@inventoryRouter.get("/reports/summary")
def summary_report(school_id: str):
    assets = Asset.objects(school_id=school_id, is_deleted=False)
    summary = {}
    for a in assets:
        cat = a.category.name if a.category else "Uncategorized"
        summary[cat] = summary.get(cat, 0) + a.quantity
    return summary

@inventoryRouter.get("/reports/lost")
def lost_or_damaged(school_id: str):
    items = Asset.objects(school_id=school_id, condition__in=["Damaged", "Lost"], is_deleted=False)
    return [{"name": a.name, "condition": a.condition, "location": a.location} for a in items]

@inventoryRouter.get("/reports/location")
def location_wise_assets(school_id: str):
    result = {}
    assets = Asset.objects(school_id=school_id, is_deleted=False)
    for a in assets:
        loc = a.location
        result[loc] = result.get(loc, []) + [a.name]
    return result
